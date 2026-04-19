#!/usr/bin/env python3
"""
update_telemetry_free.py

Appends GitHub + PyPI + Google Trends snapshots into the workbook:
  - GitHub_Snapshots
  - PyPI_Snapshots
  - Trends_Snapshots

Usage:
  python update_telemetry_free.py quantum_platform_telemetry_model_free.xlsx

Automation:
  - Run weekly (recommended) and let Excel compute deltas.

Data sources:
  - GitHub REST API: https://docs.github.com/rest
  - PyPIStats API: https://pypistats.org/api/
  - Google Trends (free) via pytrends: https://github.com/GeneralMills/pytrends

Notes:
  - Google Trends returns *relative* values (0–100) within the queried keyword set and timeframe.
    Keep your keyword set stable across runs for meaningful deltas.
"""
import sys
import time
import requests
from datetime import datetime, timedelta, date
from openpyxl import load_workbook

GITHUB_API = "https://api.github.com"
PYPIS_API  = "https://pypistats.org/api"

def read_config(cfg_ws):
    gh_token = cfg_ws["B4"].value
    frequency_days = int(cfg_ws["B5"].value or 7)

    # Trends settings
    trends_geo = (cfg_ws["B16"].value or "").strip()
    trends_timeframe = (cfg_ws["B17"].value or "today 12-m").strip()
    trends_method = (cfg_ws["B18"].value or "weekly_avg").strip()
    anchor_kw = (cfg_ws["B19"].value or "").strip()  # optional: known-volume anchor keyword
    anchor_monthly = cfg_ws["B20"].value
    try:
        anchor_monthly = float(anchor_monthly) if anchor_monthly else None
    except Exception:
        anchor_monthly = None

    # Table starts at row 22 (after header at row 21)
    items = []
    r = 22
    while True:
        company = cfg_ws[f"A{r}"].value
        if not company:
            break
        include = (cfg_ws[f"G{r}"].value or "N").strip().upper() == "Y"
        items.append({
            "company": company,
            "type": cfg_ws[f"B{r}"].value,
            "owner": (cfg_ws[f"C{r}"].value or "").strip(),
            "repo": (cfg_ws[f"D{r}"].value or "").strip(),
            "pypi": (cfg_ws[f"E{r}"].value or "").strip(),
            "trend_kw": (cfg_ws[f"F{r}"].value or "").strip(),
            "include": include
        })
        r += 1

    return gh_token, frequency_days, trends_geo, trends_timeframe, trends_method, anchor_kw, anchor_monthly, items


def ensure_sheet(wb, name, headers=None):
    """Return worksheet; create with headers if missing."""
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    if headers:
        ws.append(headers)
    return ws


def normalize_header(h: str):
    s = (h or "").strip().lower()
    # normalize unicode minus/dash
    for ch in ["–", "—", "−"]:
        s = s.replace(ch, "-")
    # replace spaces and slashes with underscores
    s = s.replace(" ", "_").replace("/", "_")
    return s


def trends_headers(ws):
    """
    Detect existing header layout to avoid misaligned columns.
    Returns (headers, normalized_headers)
    """
    if ws.max_row >= 1:
        existing = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        norm = [normalize_header(h) for h in existing]
        if "est_monthly_searches" in norm:
            return existing, norm
        # legacy layout (7 cols)
        if set(norm) >= {"snapshot_date", "company", "trend_keyword"}:
            return existing, norm
    # default layout (10 cols with est searches + anchor info)
    headers = [
        "Snapshot Date",
        "Company",
        "Trend Keyword",
        "Trends Index (0-100)",
        "Est Monthly Searches",
        "Method",
        "Geo",
        "Timeframe",
        "Anchor Keyword",
        "Anchor Monthly",
    ]
    ws.append(headers)
    return headers, [normalize_header(h) for h in headers]


def pypi_headers(ws):
    """
    Detect existing PyPI header layout; extend with MoM/YoY columns if missing.
    """
    if ws.max_row >= 1 and any(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        existing = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        # trim trailing None cells
        while existing and existing[-1] is None:
            existing = existing[:-1]
        existing = list(existing)
        norm = [normalize_header(h) for h in existing]
        # Do NOT add new columns if headers already exist; respect user-defined columns.
        return existing, norm
    headers = [
        "Snapshot Date",
        "Company",
        "PyPI Package",
        "Downloads Last Day",
        "Downloads Last Week",
        "Downloads Last Month",
        "Full Last Month",
        "Prev Month",
        "-5 Months Ago",
        "-4 Months Ago",
        "-3 Months Ago",
        "-2 Months Ago",
    ]
    ws.append(headers)
    return headers, [normalize_header(h) for h in headers]

def gh_headers(token: str):
    h = {"Accept": "application/vnd.github+json"}
    if token and "<PASTE_TOKEN_HERE>" not in str(token):
        h["Authorization"] = f"Bearer {token}"
    return h

def gh_list_paginated(path, token, params=None, limit_pages=10):
    out = []
    params = dict(params or {})
    params.setdefault("per_page", 100)
    for page in range(1, limit_pages+1):
        params["page"] = page
        url = f"{GITHUB_API}{path}"
        r = requests.get(url, headers=gh_headers(token), params=params, timeout=60)
        r.raise_for_status()
        chunk = r.json()
        if not isinstance(chunk, list) or not chunk:
            break
        out.extend(chunk)
        if len(chunk) < params["per_page"]:
            break
        time.sleep(0.2)
    return out

def gh_get(path, token, params=None):
    url = f"{GITHUB_API}{path}"
    r = requests.get(url, headers=gh_headers(token), params=params, timeout=60)
    r.raise_for_status()
    return r.json()

def count_commits_since(owner, repo, token, since_dt: datetime):
    commits = gh_list_paginated(
        f"/repos/{owner}/{repo}/commits",
        token,
        params={"since": since_dt.isoformat() + "Z"},
        limit_pages=20,
    )
    uniq_auth = {(c.get("author") or {}).get("login") for c in commits if (c.get("author") or {}).get("login")}
    return len(commits), len(uniq_auth)

def unique_issue_authors(owner, repo, token, since_dt: datetime):
    # GitHub returns PRs as issues too; filter by pull_request field.
    issues = gh_list_paginated(
        f"/repos/{owner}/{repo}/issues",
        token,
        params={"since": since_dt.isoformat() + "Z", "state": "all"},
        limit_pages=20,
    )
    issue_auth = set()
    pr_auth = set()
    for it in issues:
        user = (it.get("user") or {}).get("login")
        if not user:
            continue
        if "pull_request" in it:
            pr_auth.add(user)
        else:
            issue_auth.add(user)
    return len(issue_auth), len(pr_auth)

def releases_last_12m(owner, repo, token):
    rels = gh_list_paginated(f"/repos/{owner}/{repo}/releases", token, limit_pages=2)
    cutoff = datetime.utcnow() - timedelta(days=365)
    c = 0
    for rel in rels:
        dt = rel.get("published_at") or rel.get("created_at")
        if not dt:
            continue
        try:
            t = datetime.fromisoformat(dt.replace("Z", "+00:00")).replace(tzinfo=None)
        except Exception:
            continue
        if t >= cutoff:
            c += 1
    return c

def pypistats_recent(package: str):
    url = f"{PYPIS_API}/packages/{package}/recent"
    delay = 1.0
    for attempt in range(8):
        try:
            r = requests.get(url, timeout=60)
            r.raise_for_status()
            j = r.json()
            d = j.get("data") or {}
            return int(d.get("last_day") or 0), int(d.get("last_week") or 0), int(d.get("last_month") or 0)
        except requests.HTTPError as e:
            # retry on 429 / 5xx
            status = getattr(e.response, "status_code", None)
            if status in (429, 500, 502, 503, 504) and attempt < 7:
                time.sleep(delay + 0.3 * attempt)
                delay = min(delay * 2, 60)
                continue
            raise
        except Exception:
            if attempt < 7:
                time.sleep(delay + 0.2 * attempt)
                delay = min(delay * 2, 60)
                continue
            raise


def pypistats_monthlies(package: str):
    """
    Return downloads for:
      - last_full_month (complete month before current)
      - prev_month (month before that)
      - m_minus5 .. m_minus2 (older months, within retention window)
    Uses PyPIStats overall endpoint (daily totals) and aggregates by month.
    """
    url = f"{PYPIS_API}/packages/{package}/overall"
    delay = 1.0
    for attempt in range(8):
        try:
            r = requests.get(url, params={"mirrors": "false"}, timeout=60)
            r.raise_for_status()
            j = r.json()
            data = j.get("data") or []
            # aggregate by month (YYYY-MM)
            monthly = {}
            for drec in data:
                if drec.get("category") != "without_mirrors":
                    continue
                dt_str = drec.get("date")
                if not dt_str:
                    continue
                try:
                    dt = date.fromisoformat(dt_str)
                except Exception:
                    continue
                key = dt.strftime("%Y-%m")
                monthly[key] = monthly.get(key, 0) + int(drec.get("downloads") or 0)

            today = date.today()
            # last full month
            last_full = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
            prev = (last_full - timedelta(days=1)).replace(day=1)
            # helper to shift months without dateutil
            def shift_months(dt: date, delta_months: int) -> date:
                y = dt.year + (dt.month - 1 + delta_months) // 12
                m = (dt.month - 1 + delta_months) % 12 + 1
                return date(y, m, 1)
            m_minus2 = shift_months(last_full, -2)
            m_minus3 = shift_months(last_full, -3)
            m_minus4 = shift_months(last_full, -4)
            m_minus5 = shift_months(last_full, -5)

            def grab(dt):
                key = dt.strftime("%Y-%m")
                return monthly.get(key)

            return (
                grab(last_full),
                grab(prev),
                grab(m_minus5),
                grab(m_minus4),
                grab(m_minus3),
                grab(m_minus2),
            )
        except requests.HTTPError as e:
            status = getattr(e.response, "status_code", None)
            if status in (429, 500, 502, 503, 504) and attempt < 7:
                time.sleep(delay + 0.3 * attempt)
                delay = min(delay * 2, 60)
                continue
            raise
        except Exception:
            if attempt < 7:
                time.sleep(delay + 0.2 * attempt)
                delay = min(delay * 2, 60)
                continue
            raise

def trends_fetch(keywords, geo, timeframe, method, anchor_kw=None, anchor_monthly=None):
    """
    Returns dict keyword -> {"index": 0-100, "est_searches": float|None} based on Google Trends.
    method:
      - weekly_avg: average over last 4 weekly points (resampled weekly)
      - latest_week: latest available weekly point
    anchor_kw + anchor_monthly let us scale relative index to an estimated monthly search count.
    """
    try:
        from pytrends.request import TrendReq
    except ImportError as e:
        raise RuntimeError("pytrends not installed. Install with: pip install pytrends") from e

    all_kws = list(dict.fromkeys([k for k in keywords if k]))
    if anchor_kw:
        all_kws.append(anchor_kw)

    pytrends = TrendReq(hl="en-US", tz=0)
    pytrends.build_payload(all_kws, cat=0, timeframe=timeframe, geo=geo, gprop="")
    df = pytrends.interest_over_time()
    if df is None or df.empty:
        return {k: {"index": None, "est_searches": None} for k in keywords}
    # Drop "isPartial" if present
    if "isPartial" in df.columns:
        df = df.drop(columns=["isPartial"])

    # Force weekly cadence before aggregating so averaging is meaningful
    df = df.resample("W").mean()
    if df.empty:
        return {k: {"index": None, "est_searches": None} for k in keywords}

    if method == "latest_week":
        s = df.iloc[-1]
    else:
        s = df.tail(4).mean(numeric_only=True)

    out = {}
    anchor_avg = None
    if anchor_kw and anchor_kw in s and anchor_monthly:
        try:
            anchor_avg = float(s[anchor_kw]) or None
        except Exception:
            anchor_avg = None

    for k in keywords:
        idx = None
        est = None
        try:
            idx = int(round(float(s[k]), 0))
        except Exception:
            idx = None
        if anchor_avg and idx is not None and anchor_avg > 0:
            est = round(anchor_monthly * (idx / anchor_avg), 1)
        out[k] = {"index": idx, "est_searches": est}
    return out

def main():
    if len(sys.argv) < 2:
        print("Usage: python update_telemetry_free.py quantum_platform_telemetry_model_free.xlsx")
        sys.exit(2)

    xlsx = sys.argv[1]
    wb = load_workbook(xlsx)
    cfg = wb["Config"]
    gh_token, freq_days, geo, timeframe, method, anchor_kw, anchor_monthly, items = read_config(cfg)

    snap_date = date.today()

    sh_gh = ensure_sheet(wb, "GitHub_Snapshots")
    sh_pypi = ensure_sheet(wb, "PyPI_Snapshots")
    pypi_headers_row, pypi_norm = pypi_headers(sh_pypi)
    sh_tr = ensure_sheet(wb, "Trends_Snapshots")
    tr_headers, tr_norm = trends_headers(sh_tr)

    # GitHub + PyPI per config row
    for it in items:
        if not it["include"]:
            continue

        company = it["company"]
        owner = it["owner"]
        repo  = it["repo"]
        pypi  = it["pypi"]

        if owner and repo:
            try:
                repo_meta = gh_get(f"/repos/{owner}/{repo}", gh_token)
                stars = repo_meta.get("stargazers_count")
                forks = repo_meta.get("forks_count")
                open_issues_total = repo_meta.get("open_issues_count")  # includes PRs
                pulls = gh_list_paginated(f"/repos/{owner}/{repo}/pulls", gh_token, params={"state": "open"}, limit_pages=1)
                open_prs = len(pulls)

                commits_7d, _ = count_commits_since(owner, repo, gh_token, datetime.utcnow() - timedelta(days=7))
                _, uniq_commit_auth_30d = count_commits_since(owner, repo, gh_token, datetime.utcnow() - timedelta(days=30))
                uniq_issue_auth_30d, uniq_pr_auth_30d = unique_issue_authors(owner, repo, gh_token, datetime.utcnow() - timedelta(days=30))
                rel_12m = releases_last_12m(owner, repo, gh_token)

                sh_gh.append([
                    snap_date, company, owner, repo,
                    stars, forks, open_issues_total, open_prs,
                    commits_7d, uniq_commit_auth_30d,
                    uniq_issue_auth_30d, uniq_pr_auth_30d,
                    rel_12m
                ])
            except Exception as e:
                sh_gh.append([
                    snap_date, company, owner, repo,
                    None, None, None, None,
                    None, None, None, None,
                    None
                ])
                print(f"[WARN] GitHub fetch failed for {owner}/{repo}: {e}")
            time.sleep(0.2)

        if pypi:
            try:
                d, w, m = pypistats_recent(pypi)
                m_full, m_prev, m_m5, m_m4, m_m3, m_m2 = pypistats_monthlies(pypi)
                row_data = {
                    "snapshot_date": snap_date,
                    "company": company,
                    "pypi_package": pypi,
                    "downloads_last_day": d,
                    "downloads_last_week": w,
                    "downloads_last_month": m,
                    "full_last_month": m_full,
                    "prev_month": m_prev,
                    "-5_months_ago": m_m5,
                    "-4_months_ago": m_m4,
                    "-3_months_ago": m_m3,
                    "-2_months_ago": m_m2,
                }
                out_row = []
                for raw, norm_key in zip(pypi_headers_row, pypi_norm):
                    val = row_data.get(norm_key)
                    if val is None and "months_ago" in norm_key:
                        txt = normalize_header(raw)
                        off = None
                        for part in txt.replace("__", "_").split("_"):
                            if part.startswith("-") and part[1:].isdigit():
                                off = int(part)
                                break
                        if off is not None:
                            val = row_data.get(f"-{abs(off)}_months_ago")
                    out_row.append(val)
                sh_pypi.append(out_row)
            except Exception as e:
                row_data = {
                    "snapshot_date": snap_date,
                    "company": company,
                    "pypi_package": pypi,
                    "downloads_last_day": None,
                    "downloads_last_week": None,
                    "downloads_last_month": None,
                    "full_last_month": None,
                    "prev_month": None,
                    "-5_months_ago": None,
                    "-4_months_ago": None,
                    "-3_months_ago": None,
                    "-2_months_ago": None,
                }
                row = [row_data.get(n) for n in pypi_norm]
                sh_pypi.append(row)
                print(f"[WARN] PyPIStats failed for {pypi}: {e}")

    # Trends: de-duplicate keywords, fetch once, then write one row per company keyword
    kws = []
    comp_kw = []  # (company, kw)
    for it in items:
        if not it["include"]:
            continue
        kw = it["trend_kw"]
        if not kw:
            continue
        comp_kw.append((it["company"], kw))
        if kw not in kws:
            kws.append(kw)

    if kws:
        try:
            vals = trends_fetch(kws, geo, timeframe, method, anchor_kw, anchor_monthly)
        except Exception as e:
            print(f"[WARN] Google Trends fetch failed: {e}")
            vals = {k: {"index": None, "est_searches": None} for k in kws}

        for company, kw in comp_kw:
            res = vals.get(kw, {})
            row_data = {
                "snapshot_date": snap_date,
                "company": company,
                "trend_keyword": kw,
                "trends_index_(0-100)": res.get("index"),
                "est_monthly_searches": res.get("est_searches"),
                "method": method,
                "geo": geo,
                "timeframe": timeframe,
                "anchor_keyword": anchor_kw or None,
                "anchor_monthly": anchor_monthly,
            }
            # build row matching detected headers
            row = []
            for h_norm in tr_norm:
                row.append(row_data.get(h_norm))
            sh_tr.append(row)

    wb.save(xlsx)
    print(f"Done. Appended snapshot(s) to: {xlsx}")

if __name__ == "__main__":
    main()
